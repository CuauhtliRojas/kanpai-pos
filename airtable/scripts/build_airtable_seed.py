"""Build a normalized Airtable seed from fixed JSON and a live Excel file."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_EXCEL = Path("airtable/imports/Kanpai.xlsx")
DEFAULT_FIXED = Path("airtable/seeds/kanpai_fixed_seed.v1.json")
DEFAULT_OUTPUT = Path("airtable/seeds/kanpai_seed.generated.json")
DEFAULT_RECIPE_MULTIPLIER = Decimal("1")
YAKITORI_PREPARATION_SKUS = (
    "YAK-COC-POLL",
    "YAK-COC-PORK",
    "YAK-COC-PUL",
    "YAK-COC_CAM",
    "YAK-COC-VER",
    "YAK-COC-HONG",
    "YAK-COC-MIX",
)

TABLE_ORDER = (
    "Roles",
    "Unidades",
    "ZonasServicio",
    "MetodosPago",
    "CanalesNotificacion",
    "EstacionesProduccion",
    "Impresoras",
    "CategoriasMenu",
    "Empleados",
    "Mesas",
    "InsumosInventario",
    "Productos",
    "GruposVarianteProducto",
    "OpcionesVarianteProducto",
    "AsignacionesProductoEstacion",
    "RecetasProducto",
)
NATURAL_KEYS = {
    "Roles": "clave_rol",
    "Unidades": "clave_unidad",
    "ZonasServicio": "clave_zona",
    "MetodosPago": "clave_metodo",
    "CanalesNotificacion": "clave_canal",
    "EstacionesProduccion": "clave_estacion",
    "Impresoras": "clave_impresora",
    "CategoriasMenu": "nombre",
    "Empleados": "codigo_empleado",
    "Mesas": "codigo_mesa",
    "InsumosInventario": "codigo_insumo",
    "Productos": "sku",
    "GruposVarianteProducto": ("producto", "nombre"),
    "OpcionesVarianteProducto": ("grupo_variante", "nombre"),
    "AsignacionesProductoEstacion": "nombre_registro",
    "RecetasProducto": "nombre_registro",
}
LINK_FIELDS = {
    "Impresoras": {"estacion": ("EstacionesProduccion", "clave_estacion")},
    "Empleados": {"roles": ("Roles", "clave_rol")},
    "Mesas": {"zona": ("ZonasServicio", "clave_zona")},
    "InsumosInventario": {"unidad_base": ("Unidades", "clave_unidad")},
    "Productos": {"categoria": ("CategoriasMenu", "nombre")},
    "GruposVarianteProducto": {
        "producto": ("Productos", "sku"),
    },
    "OpcionesVarianteProducto": {
        "grupo_variante": (
            "GruposVarianteProducto",
            ("producto", "nombre"),
        ),
        "producto_opcional": ("Productos", "sku"),
        "estacion": ("EstacionesProduccion", "clave_estacion"),
    },
    "AsignacionesProductoEstacion": {
        "producto": ("Productos", "sku"),
        "estacion": ("EstacionesProduccion", "clave_estacion"),
    },
    "RecetasProducto": {
        "producto": ("Productos", "sku"),
        "insumo": ("InsumosInventario", "codigo_insumo"),
    },
}

EXPECTED_COLUMNS = {
    "Insumos": (
        "Clave",
        "Nombre",
        "Tipo",
        "Unidad_Base",
        "Stock_Minimo",
        "Costo_Unitario",
    ),
    "Productos": (
        "SKU",
        "Nombre",
        "Variante",
        "Descripcion",
        "Precio",
        "Categoria",
        "Estacion",
        "Requiere_Produccion",
        "Activo",
        "Visible_POS",
        "Inventory_Recipe_Multiplier",
    ),
    "Combo_Opciones": (
        "Combo_SKU",
        "Grupo",
        "Min_Selecciones",
        "Max_Selecciones",
        "Permite_Repetir",
        "Opcion_SKU",
        "Opcion_Nombre",
        "Cantidad_Componente",
        "Usa_Receta_De_Opcion",
        "Nota",
    ),
    "Recetas": ("SKU_Producto", "Clave_Insumo", "Cantidad", "Merma_Pct"),
}


@dataclass(frozen=True)
class SeedIssue:
    level: str
    code: str
    message: str
    sheet: str = ""
    row: int | None = None


@dataclass
class BuildResult:
    tables: dict[str, list[dict[str, Any]]]
    issues: list[SeedIssue]
    stats: dict[str, int]
    excel_present: bool

    @property
    def warnings(self) -> list[SeedIssue]:
        return [issue for issue in self.issues if issue.level == "warning"]

    @property
    def errors(self) -> list[SeedIssue]:
        return [issue for issue in self.issues if issue.level == "error"]

    def as_json(self) -> dict[str, Any]:
        return {
            "format_version": "1.0.0",
            "excel_present": self.excel_present,
            "stats": self.stats,
            "issues": [asdict(issue) for issue in self.issues],
            "tables": self.tables,
        }


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def folded(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean_string(value))
    return "".join(char for char in text if not unicodedata.combining(char)).casefold()


def normalize_header(value: Any) -> str:
    return folded(value).replace(" ", "_")


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return value != 0
    return folded(value) in {"si", "checked", "true", "1", "yes", "x"}


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or clean_string(value) == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = clean_string(value).replace("$", "").replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def money_to_cents(value: Any) -> int | None:
    number = parse_decimal(value)
    if number is None:
        return None
    return int((number * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def quantity(value: Any) -> tuple[Decimal | None, bool]:
    number = parse_decimal(value)
    if number is not None:
        return number, False
    match = re.fullmatch(r"\s*([+-]?\d+(?:[.,]\d+)?)\s*[a-zA-ZáéíóúÁÉÍÓÚ]+\s*", str(value))
    if not match:
        return None, False
    return Decimal(match.group(1).replace(",", ".")), True


def airtable_number(value: Decimal) -> int | float:
    integral = value.to_integral_value()
    return int(integral) if value == integral else float(value)


def read_excel_rows(
    path: Path, issues: list[SeedIssue]
) -> dict[str, list[tuple[int, dict[str, Any]]]]:
    result = {sheet: [] for sheet in EXPECTED_COLUMNS}
    if not path.exists():
        issues.append(
            SeedIssue(
                "warning",
                "excel_missing",
                f"No existe {path}; se construye únicamente el seed fijo.",
            )
        )
        return result

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:  # openpyxl emits several format-specific exceptions
        issues.append(SeedIssue("error", "excel_unreadable", str(error)))
        return result

    try:
        for sheet_name, expected in EXPECTED_COLUMNS.items():
            if sheet_name not in workbook.sheetnames:
                issues.append(
                    SeedIssue(
                        "warning",
                        "sheet_missing",
                        f"Falta la hoja esperada {sheet_name}.",
                        sheet_name,
                    )
                )
                continue
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header_values = next(rows, None)
            if header_values is None:
                issues.append(
                    SeedIssue("warning", "empty_sheet", "La hoja está vacía.", sheet_name)
                )
                continue
            headers = [normalize_header(value) for value in header_values]
            expected_headers = {normalize_header(value) for value in expected}
            missing = sorted(expected_headers - set(headers))
            if missing:
                issues.append(
                    SeedIssue(
                        "warning",
                        "columns_missing",
                        f"Columnas ausentes: {', '.join(missing)}",
                        sheet_name,
                        1,
                    )
                )
            for row_number, values in enumerate(rows, start=2):
                row = dict(zip(headers, values, strict=False))
                if any(clean_string(value) for value in values):
                    result[sheet_name].append((row_number, row))
    finally:
        workbook.close()
    return result


ITEM_TYPE_ALIASES = {
    "alimento": "Alimento",
    "alimentos": "Alimento",
    "bebida": "Bebida",
    "bebidas": "Bebida",
    "desechable": "Desechable",
    "desechables": "Desechable",
    "limpieza": "Limpieza",
    "otro": "Otro",
    "otros": "Otro",
}


def normalize_item_type(raw: Any) -> tuple[str, bool]:
    item_type = folded(raw)
    if not item_type:
        return "Otro", False
    if item_type in ITEM_TYPE_ALIASES:
        return ITEM_TYPE_ALIASES[item_type], False
    return "Otro", True


CATEGORY_ALIASES = {
    "comida": "Yakitori",
    "yakitori": "Yakitori",
    "cerveza": "Cervezas",
    "cervezas": "Cervezas",
    "sake": "Sake",
    "refresco": "Refrescos",
    "refrescos": "Refrescos",
    "bebida sin alcohol": "Bebidas sin alcohol",
    "bebidas sin alcohol": "Bebidas sin alcohol",
    "bebida alcohol": "Bebidas alcohol",
    "bebidas alcohol": "Bebidas alcohol",
    "alcohol": "Bebidas alcohol",
}


def normalize_category(raw: Any, product_text: str) -> tuple[str, bool]:
    category = folded(raw)
    if category in CATEGORY_ALIASES:
        return CATEGORY_ALIASES[category], False
    if category in {"bebida", "bebidas"}:
        clues = folded(product_text)
        sku_prefix = clean_string(product_text).split("-", 1)[0].upper()
        by_prefix = {
            "SAK": "Sake",
            "CER": "Cervezas",
            "REF": "Refrescos",
            "MOC": "Bebidas sin alcohol",
            "RUS": "Bebidas sin alcohol",
            "MIX": "Bebidas alcohol",
        }
        if sku_prefix in by_prefix:
            return by_prefix[sku_prefix], False
        for token, normalized in (
            ("sake", "Sake"),
            ("cerve", "Cervezas"),
            ("refresco", "Refrescos"),
            ("coct", "Bebidas alcohol"),
        ):
            if token in clues:
                return normalized, False
        return "Bebida", True
    normalized = clean_string(raw)
    if not normalized:
        return "", True
    return normalized[0].upper() + normalized[1:].lower(), True


def normalize_station(raw: Any, category: str) -> tuple[str, bool]:
    station = folded(raw).replace(" ", "_")
    known = {
        "cocina": "COCINA",
        "barra": "BARRA",
    }
    if station in known:
        return known[station], False
    return "", bool(station)


def _add_incomplete(
    issues: list[SeedIssue], stats: Counter[str], sheet: str, row: int, message: str
) -> None:
    stats[f"{sheet.casefold()}_incomplete"] += 1
    issues.append(SeedIssue("warning", "incomplete_row", message, sheet, row))


def _dedupe(
    records: list[dict[str, Any]],
    key: str | tuple[str, ...],
    table: str,
    issues: list[SeedIssue],
) -> list[dict[str, Any]]:
    key_fields = (key,) if isinstance(key, str) else key
    seen: set[tuple[Any, ...]] = set()
    result = []
    for record in records:
        value = tuple(_seed_key_part(record.get(field)) for field in key_fields)
        if value in seen:
            issues.append(
                SeedIssue(
                    "warning",
                    "duplicate_natural_key",
                    f"Duplicado omitido: {table}.{key}={value}",
                )
            )
            continue
        seen.add(value)
        result.append(record)
    return result


def _seed_key_part(value: Any) -> Any:
    if isinstance(value, (list, tuple)):
        return tuple(_seed_key_part(item) for item in value)
    return clean_string(value)


def build_seed(excel_path: Path = DEFAULT_EXCEL, fixed_path: Path = DEFAULT_FIXED) -> BuildResult:
    issues: list[SeedIssue] = []
    stats: Counter[str] = Counter()
    fixed = json.loads(fixed_path.read_text(encoding="utf-8"))
    excel_rows = read_excel_rows(excel_path, issues)

    tables: dict[str, list[dict[str, Any]]] = {table: [] for table in TABLE_ORDER}
    tables["Roles"] = fixed["Roles"]
    tables["Unidades"] = fixed["Unidades"]
    tables["ZonasServicio"] = fixed["ZonasServicio"]
    tables["MetodosPago"] = fixed["MetodosPago"]
    tables["CanalesNotificacion"] = fixed["CanalesNotificacion"]
    tables["EstacionesProduccion"] = fixed["EstacionesProduccion"]
    tables["Impresoras"] = fixed["Impresoras"]
    tables["CategoriasMenu"] = fixed["CategoriasMenu"]
    tables["Mesas"] = [
        {
            **{
                key: value
                for key, value in row.items()
                if key not in {"estado_temporal", "zona"}
            },
            "zona": [row["zona"]],
        }
        for row in fixed["Mesas"]
    ]
    tables["Empleados"] = [
        {
            "codigo_empleado": row["codigo_empleado"],
            "nombre_completo": row["nombre_completo"],
            "alias_pos": row["alias_pos"],
            "roles": [row["rol_clave"]],
            "activo": row["activo"],
        }
        for row in fixed["Empleados"]
    ]

    insumo_keys: set[str] = set()
    warned_insumo_types: set[str] = set()
    for row_number, row in excel_rows["Insumos"]:
        code = clean_string(row.get("clave"))
        name = clean_string(row.get("nombre"))
        unit = clean_string(row.get("unidad_base")).upper().replace(" ", "_")
        if not code or not name or not unit:
            _add_incomplete(
                issues, stats, "Insumos", row_number, "Se requieren Clave, Nombre y Unidad_Base."
            )
            continue
        if code in insumo_keys:
            stats["insumos_duplicates"] += 1
            issues.append(
                SeedIssue("warning", "duplicate_natural_key", f"Clave duplicada: {code}", "Insumos", row_number)
            )
            continue
        stock = parse_decimal(row.get("stock_minimo"))
        cost = money_to_cents(row.get("costo_unitario"))
        if row.get("stock_minimo") not in (None, "") and stock is None:
            issues.append(SeedIssue("warning", "invalid_number", "Stock_Minimo inválido; se usa 0.", "Insumos", row_number))
        if row.get("costo_unitario") not in (None, "") and cost is None:
            issues.append(SeedIssue("warning", "invalid_number", "Costo_Unitario inválido; se usa 0.", "Insumos", row_number))
        raw_type = clean_string(row.get("tipo"))
        item_type, item_type_warning = normalize_item_type(raw_type)
        if item_type_warning:
            stats["insumos_type_unknown"] += 1
            normalized_type = folded(raw_type)
            if normalized_type not in warned_insumo_types:
                issues.append(
                    SeedIssue(
                        "warning",
                        "insumo_type_unknown",
                        f"Tipo {raw_type!r} no reconocido; se usa 'Otro'.",
                        "Insumos",
                        row_number,
                    )
                )
                warned_insumo_types.add(normalized_type)
        tables["InsumosInventario"].append(
            {
                "codigo_insumo": code,
                "nombre": name,
                "unidad_base": [unit],
                "tipo_insumo": item_type,
                "stock_minimo": airtable_number(stock or Decimal(0)),
                "costo_unitario_centavos": cost or 0,
                "activo": True,
            }
        )
        insumo_keys.add(code)
        stats["insumos_valid"] += 1

    product_keys: set[str] = set()
    product_requires_production: dict[str, bool] = {}
    active_visible_products: set[str] = set()
    live_categories: list[str] = []
    for row_number, row in excel_rows["Productos"]:
        sku = clean_string(row.get("sku"))
        name = clean_string(row.get("nombre"))
        price = money_to_cents(row.get("precio"))
        if not sku or not name or price is None:
            _add_incomplete(
                issues, stats, "Productos", row_number, "Se requieren SKU, Nombre y Precio numérico."
            )
            continue
        if sku in product_keys:
            stats["productos_duplicates"] += 1
            issues.append(SeedIssue("warning", "duplicate_natural_key", f"SKU duplicado: {sku}", "Productos", row_number))
            continue
        variant = clean_string(row.get("variante"))
        if folded(variant) == "no":
            variant = ""
        category, category_warning = normalize_category(
            row.get("categoria"),
            f"{sku} {name} {variant} {clean_string(row.get('descripcion'))}",
        )
        if not category:
            _add_incomplete(issues, stats, "Productos", row_number, "Categoria vacía o no resoluble.")
            continue
        if category_warning:
            issues.append(SeedIssue("warning", "category_mapping_warning", f"Categoría viva conservada como {category!r}.", "Productos", row_number))
        if category not in live_categories:
            live_categories.append(category)
        display_name = clean_string(f"{name} {variant}")
        multiplier = parse_decimal(row.get("inventory_recipe_multiplier"))
        if multiplier is None:
            multiplier = DEFAULT_RECIPE_MULTIPLIER
            issues.append(
                SeedIssue(
                    "warning",
                    "missing_recipe_multiplier",
                    f"Inventory_Recipe_Multiplier inválido o vacío; se usa {DEFAULT_RECIPE_MULTIPLIER}.",
                    "Productos",
                    row_number,
                )
            )
        tables["Productos"].append(
            {
                "sku": sku,
                "tipo_producto": "Simple",
                "nombre": name,
                "variante": variant,
                "nombre_visible": display_name,
                "categoria": [category],
                "precio_centavos": price,
                "multiplicador_receta_inventario": airtable_number(multiplier),
                "activo": parse_bool(row.get("activo")),
                "visible_pos": parse_bool(row.get("visible_pos")),
            }
        )
        requires_production = parse_bool(row.get("requiere_produccion"))
        product_requires_production[sku] = requires_production
        if parse_bool(row.get("activo")) and parse_bool(row.get("visible_pos")):
            active_visible_products.add(sku)
        station, station_warning = normalize_station(row.get("estacion"), category)
        if requires_production and station:
            tables["AsignacionesProductoEstacion"].append(
                {
                    "nombre_registro": f"{sku}|{station}",
                    "producto": [sku],
                    "estacion": [station],
                    "es_principal": True,
                    "activo": True,
                }
            )
        elif requires_production and station_warning:
            issues.append(SeedIssue("warning", "station_mapping_warning", f"No se pudo mapear Estacion={clean_string(row.get('estacion'))!r} para {sku}.", "Productos", row_number))
        product_keys.add(sku)
        stats["productos_valid"] += 1

    known_categories = {row["nombre"] for row in tables["CategoriasMenu"]}
    for category in live_categories:
        if category not in known_categories:
            tables["CategoriasMenu"].append(
                {"nombre": category, "orden": len(tables["CategoriasMenu"]) + 1, "activo": True}
            )
            known_categories.add(category)

    recipe_keys: set[str] = set()
    for row_number, row in excel_rows["Recetas"]:
        sku = clean_string(row.get("sku_producto"))
        insumo = clean_string(row.get("clave_insumo"))
        amount, was_text = quantity(row.get("cantidad"))
        if not sku or not insumo or amount is None:
            _add_incomplete(issues, stats, "Recetas", row_number, "Se requieren SKU_Producto, Clave_Insumo y Cantidad válida.")
            continue
        recipe_key = f"{sku}|{insumo}"
        if recipe_key in recipe_keys:
            stats["recetas_duplicates"] += 1
            issues.append(SeedIssue("warning", "duplicate_natural_key", f"Receta duplicada: {recipe_key}", "Recetas", row_number))
            continue
        orphan_parts = []
        if sku not in product_keys:
            orphan_parts.append(f"producto {sku}")
        if insumo not in insumo_keys:
            orphan_parts.append(f"insumo {insumo}")
        if orphan_parts:
            stats["recetas_orphan"] += 1
            issues.append(SeedIssue("warning", "orphan_reference", f"Referencia huérfana: {', '.join(orphan_parts)}; fila omitida.", "Recetas", row_number))
            continue
        if was_text:
            stats["recetas_text_quantity"] += 1
            issues.append(SeedIssue("warning", "text_quantity_normalized", f"Cantidad {clean_string(row.get('cantidad'))!r} convertida a {amount}.", "Recetas", row_number))
        waste = parse_decimal(row.get("merma_pct"))
        if row.get("merma_pct") not in (None, "") and waste is None:
            issues.append(SeedIssue("warning", "invalid_number", "Merma_Pct inválida; se usa 0.", "Recetas", row_number))
        tables["RecetasProducto"].append(
            {
                "nombre_registro": recipe_key,
                "producto": [sku],
                "insumo": [insumo],
                "cantidad_base": airtable_number(amount),
                "porcentaje_merma": airtable_number(waste or Decimal(0)),
                "activo": True,
            }
        )
        recipe_keys.add(recipe_key)
        stats["recetas_valid"] += 1


    combo_groups_seen: set[tuple[str, str]] = set()
    combo_options_seen: set[tuple[str, str, str]] = set()
    for row_number, row in excel_rows["Combo_Opciones"]:
        combo_sku = clean_string(row.get("combo_sku"))
        group_name = clean_string(row.get("grupo"))
        option_sku = clean_string(row.get("opcion_sku"))
        option_name = clean_string(row.get("opcion_nombre")) or option_sku
        min_select = parse_decimal(row.get("min_selecciones"))
        max_select = parse_decimal(row.get("max_selecciones"))
        component_qty = parse_decimal(row.get("cantidad_componente"))
        if (
            not combo_sku
            or not group_name
            or not option_sku
            or min_select is None
            or max_select is None
            or component_qty is None
        ):
            _add_incomplete(
                issues,
                stats,
                "Combo_Opciones",
                row_number,
                "Se requieren Combo_SKU, Grupo, Min_Selecciones, Max_Selecciones, Opcion_SKU y Cantidad_Componente.",
            )
            continue
        if combo_sku not in product_keys:
            issues.append(
                SeedIssue(
                    "error",
                    "combo_product_missing",
                    f"Combo_SKU inexistente en Productos: {combo_sku}",
                    "Combo_Opciones",
                    row_number,
                )
            )
            continue
        if option_sku not in product_keys:
            issues.append(
                SeedIssue(
                    "error",
                    "combo_option_missing",
                    f"Opcion_SKU inexistente en Productos: {option_sku}",
                    "Combo_Opciones",
                    row_number,
                )
            )
            continue
        group_key = (combo_sku, group_name)
        if group_key not in combo_groups_seen:
            tables["GruposVarianteProducto"].append(
                {
                    "producto": [combo_sku],
                    "nombre": group_name,
                    "seleccion_minima": int(min_select),
                    "seleccion_maxima": int(max_select),
                    "requerido": True,
                    "activo": True,
                }
            )
            combo_groups_seen.add(group_key)
            stats["combo_groups_valid"] += 1
        option_key = (combo_sku, group_name, option_sku)
        if option_key in combo_options_seen:
            issues.append(
                SeedIssue(
                    "warning",
                    "duplicate_combo_option",
                    f"Opción duplicada omitida: {combo_sku}|{group_name}|{option_sku}",
                    "Combo_Opciones",
                    row_number,
                )
            )
            continue
        tables["OpcionesVarianteProducto"].append(
            {
                "grupo_variante": [(combo_sku, group_name)],
                "producto_opcional": [option_sku],
                "nombre": option_name,
                "sku": option_sku,
                "diferencia_precio_centavos": 0,
                "estacion": [],
                "activo": parse_bool(row.get("usa_receta_de_opcion")),
            }
        )
        combo_options_seen.add(option_key)
        stats["combo_options_valid"] += 1

    for sku in YAKITORI_PREPARATION_SKUS:
        if sku not in product_keys:
            continue
        group_key = (sku, "Preparación")
        if group_key not in combo_groups_seen:
            tables["GruposVarianteProducto"].append(
                {
                    "producto": [sku],
                    "nombre": "Preparación",
                    "seleccion_minima": 1,
                    "seleccion_maxima": 1,
                    "requerido": True,
                    "activo": True,
                }
            )
            combo_groups_seen.add(group_key)
        for option_name in ("Tempura", "Asada"):
            tables["OpcionesVarianteProducto"].append(
                {
                    "grupo_variante": [(sku, "Preparación")],
                    "producto_opcional": [],
                    "nombre": option_name,
                    "sku": "",
                    "diferencia_precio_centavos": 0,
                    "estacion": [],
                    "activo": True,
                }
            )

    products_with_recipe = {
        record["producto"][0] for record in tables["RecetasProducto"]
    }
    combo_product_skus = {
        record["producto"][0] for record in tables["GruposVarianteProducto"]
    }
    for sku in sorted((active_visible_products - products_with_recipe) - combo_product_skus):
        level = "error" if product_requires_production.get(sku) else "warning"
        code = (
            "prepared_product_without_recipe"
            if level == "error"
            else "direct_product_without_recipe"
        )
        issues.append(
            SeedIssue(
                level,
                code,
                f"Producto visible {sku} sin receta; no se inventa consumo de inventario.",
                "Productos",
            )
        )
        stats["productos_sin_receta"] += 1
        stats[f"productos_sin_receta_{level}"] += 1

    for table, key in NATURAL_KEYS.items():
        tables[table] = _dedupe(tables[table], key, table, issues)
    for name in (
        "insumos_valid", "insumos_incomplete", "insumos_duplicates", "insumos_type_unknown",
        "productos_valid", "productos_incomplete", "productos_duplicates",
        "recetas_valid", "recetas_incomplete", "recetas_duplicates",
        "recetas_orphan", "recetas_text_quantity",
        "productos_sin_receta", "productos_sin_receta_error",
        "productos_sin_receta_warning", "combo_groups_valid",
        "combo_options_valid", "combo_opciones_incomplete",
    ):
        stats[name] += 0
    return BuildResult(tables, issues, dict(sorted(stats.items())), excel_path.exists())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--fixed", type=Path, default=DEFAULT_FIXED)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = build_seed(args.excel, args.fixed)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(result.as_json(), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Excel presente: {result.excel_present}")
    print(f"Warnings: {len(result.warnings)}")
    print(f"Errores: {len(result.errors)}")
    print(f"Seed generado: {args.output}")
    return 1 if result.errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
