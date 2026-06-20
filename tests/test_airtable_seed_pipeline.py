from pathlib import Path
import sys

from openpyxl import Workbook

SCRIPTS_DIR = Path("airtable/scripts").resolve()
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from airtable_records_client import AirtableRecordsClient, batched  # noqa: E402
from build_airtable_seed import (  # noqa: E402
    DEFAULT_FIXED,
    TABLE_ORDER,
    build_seed,
    parse_bool,
    quantity,
)
from seed_airtable_from_excel import dry_run_plan, resolve_links  # noqa: E402


class FakeAirtableClient(AirtableRecordsClient):
    def __init__(self, records_by_table):
        super().__init__("app-test", "token-test")
        self.records_by_table = records_by_table

    def list_records(self, table, *, fields=None):
        return self.records_by_table.get(table, [])


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    insumos = workbook.active
    insumos.title = "Insumos"
    insumos.append(
        ["Clave", "Nombre", "Tipo", "Unidad_Base", "Stock_Minimo", "Costo_Unitario"]
    )
    insumos.append([" INS-1 ", " Limón ", "Alimento", " pz ", 2, 3.5])
    insumos.append(["", "Incompleto", "", "", "", ""])

    productos = workbook.create_sheet("Productos")
    productos.append(
        [
            "SKU",
            "Nombre",
            "Variante",
            "Descripcion",
            "Precio",
            "Categoria",
            "Estacion",
            "Requiere_Produccion",
            "Activo",
            "Visible_POS",
        ]
    )
    productos.append(
        [
            " MIX-1   ",
            "Coctel",
            "",
            "mezcal",
            100,
            "Bebida ",
            "Barra",
            "SI",
            "checked",
            1,
        ]
    )

    recetas = workbook.create_sheet("Recetas")
    recetas.append(["SKU_Producto", "Clave_Insumo", "Cantidad", "Merma_Pct"])
    recetas.append(["MIX-1", "INS-1", "1 pz", 5])
    recetas.append(["NO-EXISTE", "INS-1", 1, 0])
    recetas.append(["", "INS-1", 1, 0])
    workbook.save(path)


def test_build_seed_normalizes_and_skips_incomplete_rows(tmp_path: Path) -> None:
    excel = tmp_path / "seed.xlsx"
    _write_workbook(excel)

    result = build_seed(excel, DEFAULT_FIXED)

    assert not result.errors
    assert result.stats["insumos_valid"] == 1
    assert result.stats["insumos_incomplete"] == 1
    assert result.stats["productos_valid"] == 1
    assert result.stats["recetas_valid"] == 1
    assert result.stats["recetas_orphan"] == 1
    assert any(issue.code == "incomplete_row" for issue in result.warnings)
    product = result.tables["Productos"][0]
    assert product["sku"] == "MIX-1"
    assert product["categoria"] == ["Bebidas alcohol"]
    assert product["activo"] is True
    assert result.tables["RecetasProducto"][0]["cantidad_base"] == 1
    assert result.tables["AsignacionesProductoEstacion"][0]["estacion"] == [
        "COCTELERIA"
    ]


def test_missing_excel_builds_fixed_seed_only(tmp_path: Path) -> None:
    result = build_seed(tmp_path / "missing.xlsx", DEFAULT_FIXED)

    assert result.excel_present is False
    assert len(result.tables["Mesas"]) == 17
    assert result.tables["Mesas"][0]["codigo_mesa"] == "M01"
    assert result.tables["Mesas"][0]["zona"] == ["BARRA"]
    assert not result.tables["Productos"]
    assert any(issue.code == "excel_missing" for issue in result.warnings)


def test_normalizers_accept_required_boolean_and_quantity_forms() -> None:
    for value in ("SI", "checked", "true", "1", 1, True):
        assert parse_bool(value) is True
    for value in ("NO", "", "false", "0", 0, False, None):
        assert parse_bool(value) is False
    assert quantity("1pz")[0] == 1
    assert quantity("1 pz")[1] is True


def test_link_resolution_uses_ids_and_never_invents_internal_links() -> None:
    issues = []
    records = [{"codigo_mesa": "M01", "zona": ["BARRA"], "activo": True}]
    indexes = {"ZonasServicio": {"BARRA": {"id": "rec-zone"}}}

    resolved = resolve_links("Mesas", records, indexes, issues)
    missing = resolve_links("Mesas", records, {"ZonasServicio": {}}, issues)

    assert resolved[0]["zona"] == ["rec-zone"]
    assert missing == []
    assert any(issue.code == "unresolved_link" for issue in issues)


def test_airtable_batches_are_capped_at_ten() -> None:
    groups = list(batched(list(range(23))))
    assert [len(group) for group in groups] == [10, 10, 3]


def test_equivalent_airtable_values_do_not_generate_updates() -> None:
    client = FakeAirtableClient({})
    existing = {
        "ITEM-1": {
            "id": "rec-item",
            "fields": {
                "code": "ITEM-1 ",
                "name": " Item ",
                "quantity": 1.0,
                "active": True,
            },
        }
    }

    plan = client.plan_upsert(
        "Items",
        "code",
        [
            {
                "code": "ITEM-1",
                "name": "Item",
                "quantity": 1,
                "active": True,
            }
        ],
        existing=existing,
    )

    assert not plan["updates"]
    assert len(plan["unchanged"]) == 1


def test_equivalent_linked_records_and_empty_fields_do_not_generate_updates() -> None:
    client = FakeAirtableClient({})
    existing = {
        "ITEM-1": {
            "id": "rec-item",
            "fields": {
                "code": "ITEM-1",
                "links": ["rec-b", "rec-a"],
            },
        }
    }

    plan = client.plan_upsert(
        "Items",
        "code",
        [
            {
                "code": "ITEM-1",
                "links": ["rec-a", "rec-b"],
                "optional_text": "",
                "optional_links": [],
                "active": False,
            }
        ],
        linked_fields={"links", "optional_links"},
        existing=existing,
    )

    assert not plan["updates"]
    assert len(plan["unchanged"]) == 1


def test_excluded_readonly_field_does_not_generate_update_or_payload() -> None:
    client = FakeAirtableClient({})
    existing = {
        "ITEM-1": {
            "id": "rec-item",
            "fields": {"code": "ITEM-1", "formula": "remote value"},
        }
    }

    plan = client.plan_upsert(
        "Items",
        "code",
        [{"code": "ITEM-1", "formula": "local value"}],
        excluded_fields={"formula"},
        existing=existing,
    )

    assert not plan["updates"]
    assert len(plan["unchanged"]) == 1


def test_mesas_dry_run_matches_by_codigo_mesa(tmp_path: Path) -> None:
    result = build_seed(tmp_path / "missing.xlsx", DEFAULT_FIXED)
    zone_records = [
        {"id": "rec-barra", "fields": {"clave_zona": "BARRA"}},
        {"id": "rec-salon", "fields": {"clave_zona": "SALON"}},
    ]
    mesa_records = []
    for mesa in result.tables["Mesas"]:
        zone_id = "rec-barra" if mesa["zona"] == ["BARRA"] else "rec-salon"
        mesa_records.append(
            {
                "id": f"rec-{mesa['codigo_mesa']}",
                "fields": {**mesa, "zona": [zone_id]},
            }
        )
    client = FakeAirtableClient(
        {
            "ZonasServicio": zone_records,
            "Mesas": mesa_records,
        }
    )
    issues = []

    summary = dry_run_plan(client, result, issues)

    assert summary["Mesas"] == {
        "creates": 0,
        "updates": 0,
        "unchanged": 17,
        "upserts": 17,
        "skipped": 0,
    }


def test_fixed_seed_contains_role_and_unit_dependencies(tmp_path: Path) -> None:
    result = build_seed(tmp_path / "missing.xlsx", DEFAULT_FIXED)

    assert TABLE_ORDER[:2] == ("Roles", "Unidades")
    assert {role["clave_rol"] for role in result.tables["Roles"]} == {
        "ADMIN",
        "GERENTE",
        "CAJERO",
        "ALMACEN",
    }
    assert {unit["clave_unidad"] for unit in result.tables["Unidades"]} == {
        "G",
        "KG",
        "LT",
        "ML",
        "OZ",
        "PZA",
    }
